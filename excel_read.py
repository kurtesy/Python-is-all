import openpyxl

wb = openpyxl.load_workbook(filename = 'ETRM_trim.xlsx')
wb1 = openpyxl.Workbook()
ws = wb['Sheet1']
dest_filename = 'etrm_struct.xlsx'
k=1
ws2 = wb1.create_sheet(title="New")
for i in range(1,100):
    if(ws.cell(row = i, column = 1).value=="Foreign Keys "):
        Head=ws.cell(row = i-1, column = 1).value
        ws2 = wb1.create_sheet(title=Head)
    if(ws.cell(row = i, column = 1).value=="Primary Key Table "):
        x=i
        ws2['A'+str(k)]="Primary Key Table"
        k+=1
    while(i>x):
        if(ws.cell(row = i, column = 1).value=="QuickCodes Columns"):
            break
        ws2['A'+str(k)]=ws.cell(row = i, column = 1).value
        k+=1
    if(ws.cell(row = i, column = 1).value=="Primary Key Column"):
        x=i
        ws2['B'+str(k)]="Primary Key Column"
        k+=1
    while(i>x):
        if(ws.cell(row = i, column = 1).value=="Foreign Key Column"):
            break
        ws2['B'+str(k)]=ws.cell(row = i, column = 1).value
        k+=1
    if(ws.cell(row = i, column = 1).value=="Foreign Key Column"):
        x=i
        ws2['C'+str(k)]="Primary Key Column"
        k+=1
    while(i>x):
        if(ws.cell(row = i, column = 1).value=="QuickCodes Table"):
            break
        ws2['C'+str(k)]=ws.cell(row = i, column = 1).value
        k+=1
wb1.save(filename = dest_filename)
